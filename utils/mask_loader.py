from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import re

import numpy as np
from openpyxl import load_workbook


PRE_SURGERY_MASK_SHEET = re.compile(
    r"^(?P<network>WB|DMN|SMN|VN)_pre_(?P<descriptor>.+)$"
)
TOPOLOGY_MASK_SHEET = re.compile(
    r"^(?P<network>WB|DMN|SMN|VN)_pre_(?!FC$).+"
)


@dataclass(frozen=True)
class FcMask:
    edges: tuple[tuple[int, int], ...]
    statistics: tuple[float, ...]


@dataclass(frozen=True)
class TopologyMaskValues:
    network: str
    sheet_name: str
    dNCR_values: tuple[float, ...]
    non_dNCR_values: tuple[float, ...]


@dataclass(frozen=True)
class PreSurgeryMaskSheet:
    network: str
    sheet_name: str
    feature_family: str


@lru_cache(maxsize=None)
def discover_pre_surgery_mask_sheets(
    workbook_path: str | Path,
) -> tuple[PreSurgeryMaskSheet, ...]:
    """Discover FC and topology masks from workbook sheet names."""
    path = Path(workbook_path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Mask workbook does not exist: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        masks = []
        seen = set()
        for sheet_name in workbook.sheetnames:
            match = PRE_SURGERY_MASK_SHEET.fullmatch(sheet_name)
            if match is None:
                continue
            network = match.group("network")
            feature_family = (
                "FC"
                if match.group("descriptor").casefold() == "fc"
                else "topology"
            )
            key = (network, feature_family)
            if key in seen:
                raise ValueError(
                    f"{path}: multiple pre-surgery {feature_family} "
                    f"mask sheets for {network}"
                )
            masks.append(
                PreSurgeryMaskSheet(
                    network=network,
                    sheet_name=sheet_name,
                    feature_family=feature_family,
                )
            )
            seen.add(key)
    finally:
        workbook.close()

    if not masks:
        raise ValueError(f"{path}: no pre-surgery mask sheets found")
    return tuple(masks)


def _numeric_column(worksheet, column: int) -> tuple[float, ...]:
    values = []
    for row in range(1, worksheet.max_row + 1):
        raw_value = worksheet.cell(row=row, column=column).value
        if isinstance(raw_value, bool) or not isinstance(
            raw_value, (int, float)
        ):
            continue
        value = float(raw_value)
        if not np.isfinite(value):
            raise ValueError(
                f"{worksheet.title}: non-finite value at "
                f"row {row}, column {column}"
            )
        values.append(value)
    return tuple(values)


@lru_cache(maxsize=None)
def load_topology_mask_values(
    workbook_path: str | Path,
) -> tuple[TopologyMaskValues, ...]:
    """Discover pre-surgery topology masks and their group values."""
    path = Path(workbook_path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Mask workbook does not exist: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        masks = []
        seen_networks = set()
        for sheet_name in workbook.sheetnames:
            match = TOPOLOGY_MASK_SHEET.fullmatch(sheet_name)
            if match is None:
                continue
            network = match.group("network")
            if network in seen_networks:
                raise ValueError(
                    f"{path}: multiple pre-surgery topology sheets for "
                    f"{network}"
                )
            worksheet = workbook[sheet_name]
            dNCR_values = _numeric_column(worksheet, 2)
            non_dNCR_values = _numeric_column(worksheet, 4)
            if not dNCR_values or not non_dNCR_values:
                raise ValueError(
                    f"{path}/{sheet_name}: topology group values must "
                    "be stored in columns B and D"
                )
            masks.append(
                TopologyMaskValues(
                    network=network,
                    sheet_name=sheet_name,
                    dNCR_values=dNCR_values,
                    non_dNCR_values=non_dNCR_values,
                )
            )
            seen_networks.add(network)
    finally:
        workbook.close()

    if not masks:
        raise ValueError(f"{path}: no pre-surgery topology sheets found")
    return tuple(masks)


@lru_cache(maxsize=None)
def load_fc_mask(
    workbook_path: str | Path,
    network: str,
    node_count: int,
) -> FcMask:
    """Load a network FC mask from its nonzero upper triangle."""
    path = Path(workbook_path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Mask workbook does not exist: {path}")
    if node_count <= 1:
        raise ValueError(
            f"Invalid node count for {network}: {node_count}"
        )

    matching_sheets = [
        mask.sheet_name
        for mask in discover_pre_surgery_mask_sheets(path)
        if mask.network == network and mask.feature_family == "FC"
    ]
    if len(matching_sheets) != 1:
        raise ValueError(
            f"{path}: expected one pre-surgery FC mask for {network}, "
            f"found {len(matching_sheets)}"
        )
    sheet_name = matching_sheets[0]

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        if (
            worksheet.max_row != node_count
            or worksheet.max_column != node_count
        ):
            raise ValueError(
                f"{path}/{sheet_name}: expected a "
                f"{node_count}x{node_count} matrix, found "
                f"{worksheet.max_row}x{worksheet.max_column}"
            )

        matrix = np.empty((node_count, node_count), dtype=float)
        for row in range(node_count):
            for column in range(node_count):
                raw_value = worksheet.cell(
                    row=row + 1,
                    column=column + 1,
                ).value
                value = 0.0 if raw_value is None else float(raw_value)
                if not np.isfinite(value):
                    raise ValueError(
                        f"{path}/{sheet_name}: non-finite value "
                        f"at row {row + 1}, column {column + 1}"
                    )
                matrix[row, column] = value
    finally:
        workbook.close()

    if not np.allclose(matrix, matrix.T, rtol=0.0, atol=1e-12):
        raise ValueError(
            f"{path}/{sheet_name}: mask matrix is not symmetric"
        )
    if not np.allclose(np.diag(matrix), 0.0, rtol=0.0, atol=1e-12):
        raise ValueError(
            f"{path}/{sheet_name}: mask diagonal must be zero"
        )

    row_indices, column_indices = np.triu_indices(node_count, k=1)
    upper_values = matrix[row_indices, column_indices]
    nonzero = upper_values != 0.0
    edges = tuple(
        (int(row), int(column))
        for row, column in zip(
            row_indices[nonzero],
            column_indices[nonzero],
        )
    )
    statistics = tuple(float(value) for value in upper_values[nonzero])
    if not edges:
        raise ValueError(
            f"{path}/{sheet_name}: no nonzero mask edges found"
        )
    if len(edges) != len(statistics):
        raise RuntimeError("FC mask edge/statistic count mismatch")

    return FcMask(
        edges=edges,
        statistics=statistics,
    )
